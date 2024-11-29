'''
This script is purpose built for an organization that needs to export their membership data out of the
Paid Memberships Pro plugin in Wordpress, divide the data up per club into individual Excel spreadsheets,
and then save them to each club's Google Drive folder. It also saves a copy of the full members list and 
orders spreadsheets to Google Drive as well.

config.yaml, client_secrets.json, and gdrive_auth.txt files are prerequisites. See the example files for details.
Logs are saved to ./output.log
'''

import pandas as pd
import yaml
import os
import io
import logging
import requests
import pydrive.auth
import pydrive.drive
from urllib3.util import Retry
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

def create_logger(working_dir):
    filepath = f'{working_dir}/output.log'
    formatter = logging.Formatter(fmt='%(asctime)s %(levelname)-8s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    handler = logging.FileHandler(filepath, mode='w')
    handler.setFormatter(formatter)
    log = logging.getLogger(__name__)
    log.setLevel(logging.DEBUG)
    log.addHandler(handler)

    return log

class Config():
    def __init__(self, filepath):
        # Load config.yaml file
        try:
            with open(filepath, 'r') as file:
                data = yaml.safe_load(file)
        except:
            logger.exception('filed to load config.yaml')
            os._exit(0)

        # Unpack data into the class object
        self.working_dir = f'{os.path.dirname(filepath)}'
        self.header = data['wordpress']['header']
        self.wp_login = data['wordpress']['wp_login']
        self.wp_admin = data['wordpress']['wp_admin']
        self.username = data['wordpress']['username']
        self.password = data['wordpress']['password']
        self.membership_url = data['wordpress']['membership_url']
        self.orders_url = data['wordpress']['orders_url']
        self.club_folders_id = data['google_drive']['club_folders_id']
        self.reports_folder_id = data['google_drive']['reports_folder_id']

    def path(self, filename):
        return f'{self.working_dir}/{filename}'

class Wordpress():
    def __init__(self, config):
        login_data = {
            'log':config.username, 
            'pwd':config.password, 
            'wp-submit':'Log In', 
            'redirect_to':config.wp_admin, 
            "rememberme": "forever", 
            'testcookie':'1' 
        }
        retry = Retry(
            total=5,
            backoff_factor=2,
            status_forcelist=[408, 429, 500, 502, 503, 504],
        )

        # Build HTTP Client
        adapter = requests.adapters.HTTPAdapter(max_retries=retry)
        client = requests.Session()
        client.headers.update(config.header)
        client.mount('https://', adapter)
        self.client =  client

        # Authenticate with Wordpress server
        login = self.http_post(config.wp_login, 'successfully logged into Wordpress', data=login_data)

        # Get membership data
        self.membership(config)

        # Get club list from membership data
        self.club_list()

        # Get orders data
        self.orders(config)

        # Close the connection
        self.client.close()

    def http_post(self, url, success_message, **kwargs):
        if 'data' in kwargs.keys():
            data = kwargs['data']
        else:
            data = {}

        try:
            response = self.client.post(url, data=data, allow_redirects=True, timeout=180)
            if response.ok:
                logger.info(success_message)
        except:
            logger.exception(f'failed POST to url: {url}')
            self.client.close()
            os._exit(0)

        return response
    
    def membership(self, config):
        # Download the membership CSV file from the Paid Membership Pro plugin
        membership_response = self.http_post(config.membership_url, 'successfully downloaded membership CSV file')
        
        # Decode and create Membership CSV
        membership_raw_data = self.decode(membership_response)
        membership_data = self.create_csv(membership_raw_data)
        logger.info(f'{len(membership_data)} members found in data')

        self.membership = membership_data
    
    def orders(self, config):
        # Download orders CSV file from the Paid Membership Pro plugin
        orders_response = self.http_post(config.orders_url, 'successfully downloaded orders CSV file')

        # Decode and create Orders CSV
        orders_raw_data = self.decode(orders_response)
        orders_data = self.create_csv(orders_raw_data)

        self.orders = orders_data

    def club_list(self):
        # Get unique club names from membership list
        try:
            self.club_list = set(self.membership['home_club'])
            logger.info(f'{len(self.club_list)} clubs found in data')
        except:
            logger.exception('failed to get club list from csv data')
    
    def decode(self, response):
        try:
            return response.content.decode('utf8')
        except:
            logger.exception(f'failed to decode response data')

    def create_csv(self, data):
        try:
            return pd.read_csv(io.StringIO(data))
        except:
            logger.exception(f'failed to create Pandas dataframe from CSV data')


class GoogleDrive():
    def __init__(self, config):
        gauth = pydrive.auth.GoogleAuth()

        # Load saved client credentials (does not exist when running this for the first time)
        gauth.LoadCredentialsFile(f'{config.path('gdrive_auth.txt')}')

        if gauth.credentials is None:
            # Authenticate with Google Drive if credentials are missing
            gauth.settings.update({
                'get_refresh_token': True, 
                'client_config_file': f'{config.path('client_secrets.json')}'
            })

            try:
                gauth.LocalWebserverAuth()
                logger.info('authenticated with Google Cloud API')
            except:
                logger.exception('failed to authenticate with Google Cloud API')
                os._exit(0)

        elif gauth.access_token_expired:

            try:
                gauth.Refresh()
                logger.info('refreshed the Google Cloud API token')
            except:
                logger.exception('failed to refresh the Google Cloud API token')
                os._exit(0)
        else:

            try:
                gauth.Authorize()
                logger.info('loaded the cached Google Cloud API token')
            except:
                logger.exception('failed to load the cached Google Cloud API token')
                os._exit(0)

        # Save the current credentials to file for the next time the script is ran
        try:
            gauth.SaveCredentialsFile(f'{config.path('gdrive_auth.txt')}')
        except:
            logger.exception('failed to save gdrive credentials file')

        self.client = pydrive.drive.GoogleDrive(gauth)

    def list_contents(self, id):
        try:
            return self.client.ListFile({'q': f"'{id}' in parents and trashed=false"}).GetList()
        except:
            logger.exception(f'failed to get list of items from Google Drive for folder id: {id}')
    
    def create_file(self, path, filename, parent, id):

        if id:
            file_metadata = {
                'title': filename,
                'id': id,
                'parents': [{'id': parent}]
            }
        else:
            logger.info(f'no club spreadsheet found. Creating...')
            # Leave ID out of the metadata so that a new file is created
            file_metadata = {
                'title': filename,
                'parents': parent
            }

        # Create a Google Drive File with metadata
        file = self.client.CreateFile(file_metadata)

        # Add content from the Excel file
        file.SetContentFile(path)

        return file

    def create_folder(self, club, id):
        folder_metadata = {
            'title': club,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': id
        }
        folder = self.client.CreateFile(folder_metadata)
        folder = self.upload(folder, f'{club} folder successfully created')
        club_folder_id = folder['id']
        logger.info(f'club folder id: {club_folder_id}')

        return club_folder_id
    
    def upload(self, file, message):
        try:
            file.Upload()
            logger.info(message)
            return file
        except:
            logger.exception(f'failed to upload item to Google Drive')

def create_spreadsheets(config, wordpress):

    def make_xlsx(filename, sheet_name, data, **kwargs):
        # Write the data to an Excel file in the current directory
        try:
            data.to_excel(filename, sheet_name=sheet_name, index=False)
        except:
            logger.exception(f'failed to export {filename} spreadsheet to Excel file')

        # Load the recently created excel spreadsheet
        wb = load_workbook(filename)
        ws = wb[sheet_name]

        # Remove unneeded columns
        if 'delete' in kwargs.keys():
            ws.delete_cols(*kwargs['delete'])

        # Format the spreadsheet as a table
        table = Table(displayName='Table1', ref='A1:' + get_column_letter(ws.max_column) + str(ws.max_row))
        ws.add_table(table)

        # Autofit column width  
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 6
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save changes
        wb.save(filename)

    # Create membership spreadsheet
    make_xlsx(f'{config.path('members_list.xlsx')}', 'members', wordpress.membership)

    # Create orders spreadsheet
    make_xlsx(f'{config.path('orders.xlsx')}', 'orders', wordpress.orders)

    # Create members spreadsheets for each club
    for club in wordpress.club_list:
        # Filter out 'unknown' home_club
        if 'unknown' not in club.lower():
            # Save to Excel
            filename = f'{config.path(f"{club}.xlsx")}'
            club_data = wordpress.membership[wordpress.membership['home_club'] == club]

            make_xlsx(filename, 'members', club_data, delete=(7, 7))


def data_upload(gdrive, wordpress, config):

    # Get contents of Google Drive reports folder
    file_list = gdrive.list_contents(config.reports_folder_id)

    # Get Google Drive file IDs for members_list.xlsx and orders.xlsx if they exist
    members_list_id = ''
    orders_id = ''
    for file in file_list:
        match file['title']:
            case 'members_list.xlsx':
                members_list_id = file['id']
            case 'orders.xlsx':
                orders_id = file['id']


    # Create and upload members_list.xlsx file
    filename = "members_list.xlsx"
    filepath = f'{config.path(filename)}'
    file = gdrive.create_file(filepath, filename, config.reports_folder_id, members_list_id)
    file = gdrive.upload(file, f'{filename} successfully uploaded')

    # Clean up local copy of the club members spreadsheet
    os.remove(filepath)

    # Create and upload members_list.xlsx file
    filename = "orders.xlsx"
    filepath = f'{config.path(filename)}'
    file = gdrive.create_file(filepath, filename, config.reports_folder_id, orders_id)
    file = gdrive.upload(file, f'{filename} successfully uploaded')

    # Clean up local copy of the club members spreadsheet
    os.remove(filepath)

    # Get contents of Google Drive clubs folder
    folder_list = gdrive.list_contents(config.club_folders_id)
    
    # Get Google Drive folder IDs for each club folder
    for club in wordpress.club_list:
        if 'unknown' not in club.lower():
            logger.info(f'club name: {club}')
            club_folder_id = ''
            for club_folder in folder_list:
                if club_folder['title'] == club:
                    club_folder_id = club_folder['id']
                    logger.info(f'club folder id: {club_folder_id}')
                    break
            
            if not club_folder_id:
                logger.info('no club folder found. Creating...')

                # Create the club folder
                club_folder_id = gdrive.create_folder(club, config.club_folders_id)
                logger.info(f'club folder id: {club_folder_id}')

            # Get a list of files within the club's Google Drive folder
            file_list = gdrive.list_contents(club_folder_id)

            # Get existing file ID if present so that we can overwrite the old spreadsheet with the new one
            filename = f'{club}.xlsx'
            gfile_id = ''
            for gfile in file_list:
                if gfile['title'] == filename:
                    gfile_id = gfile['id']
                    logger.info(f'club spreadsheet id: {gfile_id}')
                    break

            # Create and upload members_list.xlsx
            filepath = f'{config.path(filename)}'
            file = gdrive.create_file(filepath, filename, club_folder_id, gfile_id)
            file = gdrive.upload(file, f'{filename} successfully uploaded')
            
            # Clean up local copy of the club members spreadsheet
            os.remove(filepath)

def main():
    working_dir = f'{os.path.dirname(__file__)}'
    global logger 
    logger = create_logger(working_dir)
    config = Config(f'{working_dir}/config.yaml')
    wordpress = Wordpress(config)
    create_spreadsheets(config, wordpress)
    google_drive = GoogleDrive(config)
    data_upload(google_drive, wordpress, config)

if __name__ == "__main__":
    main()
